"""
build_dashboard.py
==================
Reads ops_finance.db → builds a fully formatted Excel KPI dashboard
with charts, color-coded cells, and management commentary.

Usage:
    pip install pandas openpyxl
    python build_dashboard.py
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00

# ── Colors ───────────────────────────────────────────────────────
C_DARK      = "1A1A2E"
C_ORANGE    = "E8590C"
C_BLUE      = "2A6496"
C_GREEN     = "2D7A4F"
C_YELLOW    = "E8B84B"
C_LIGHT_BG  = "F7F4EF"
C_HEADER_BG = "1A1A2E"
C_WHITE     = "FFFFFF"
C_RED_LIGHT = "FFE0D6"
C_GRN_LIGHT = "D6F0E0"
C_YLW_LIGHT = "FFF6D6"
C_BORDER    = "D4CFC3"

def side(color=C_BORDER): return Side(style="thin", color=color)
def border(all=True):
    s = side()
    return Border(left=s, right=s, top=s, bottom=s) if all else Border(bottom=side())

def hfill(hex): return PatternFill("solid", fgColor=hex)
def hfont(hex, sz=10, bold=False, italic=False):
    return Font(name="Arial", color=hex, size=sz, bold=bold, italic=italic)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

# ── Fetch data from SQLite ────────────────────────────────────────
conn = sqlite3.connect("ops_finance.db")

master = pd.read_sql("""
    SELECT week_num, week_start,
        SUM(revenue)          AS revenue,
        SUM(budget_revenue)   AS budget,
        SUM(gross_profit)     AS gross_profit,
        SUM(ebitda)           AS ebitda,
        SUM(opex)             AS opex,
        SUM(orders)           AS total_orders,
        SUM(fulfillment_cost) AS fulfillment_cost,
        AVG(headcount)        AS avg_headcount
    FROM weekly_ops GROUP BY week_num, week_start ORDER BY week_num
""", conn)

region_df = pd.read_sql("""
    WITH latest AS (SELECT MAX(week_num) AS mx FROM weekly_ops)
    SELECT region, SUM(revenue) AS revenue,
        ROUND(AVG(gpm_pct),2) AS gpm_pct,
        ROUND(AVG(ebitda_margin_pct),2) AS ebitda_pct,
        SUM(orders) AS orders,
        ROUND(AVG(cost_per_order),2) AS cost_per_order,
        ROUND(AVG(budget_var_pct),2) AS budget_var
    FROM weekly_ops GROUP BY region ORDER BY revenue DESC
""", conn)

cat_df = pd.read_sql("""
    SELECT category, SUM(orders) AS orders, SUM(returns) AS returns,
        ROUND(SUM(returns)*100.0/SUM(orders),2) AS return_rate_pct,
        SUM(revenue) AS revenue
    FROM product_data GROUP BY category ORDER BY return_rate_pct DESC
""", conn)
conn.close()

# Derived columns
master["gpm_pct"]        = (master["gross_profit"] / master["revenue"] * 100).round(2)
master["ebitda_pct"]     = (master["ebitda"]        / master["revenue"] * 100).round(2)
master["opex_pct"]       = (master["opex"]          / master["revenue"] * 100).round(2)
master["budget_var_pct"] = ((master["revenue"] - master["budget"]) / master["budget"] * 100).round(2)
master["cpo"]            = (master["fulfillment_cost"] / master["total_orders"]).round(2)
master["rev_per_hc"]     = (master["revenue"] / master["avg_headcount"]).round(0)

wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE DASHBOARD
# ════════════════════════════════════════════════════════════════
dash = wb.create_sheet("📊 Dashboard")
dash.sheet_view.showGridLines = False
dash.sheet_view.zoomScale = 90

# Row heights
for r in range(1, 60): dash.row_dimensions[r].height = 18
dash.row_dimensions[1].height = 8
dash.row_dimensions[2].height = 38
dash.row_dimensions[3].height = 8
dash.row_dimensions[4].height = 20
dash.row_dimensions[5].height = 8

# Column widths
cols_w = {"A":2,"B":18,"C":16,"D":16,"E":16,"F":16,"G":16,"H":16,"I":2}
for col, w in cols_w.items(): dash.column_dimensions[col].width = w

# ── Title bar ──────────────────────────────────────────────────
dash.merge_cells("B2:H2")
c = dash["B2"]
c.value = "INDIA OPERATIONS  ·  WEEKLY KPI DASHBOARD"
c.font  = Font(name="Arial", color=C_WHITE, size=16, bold=True)
c.fill  = hfill(C_DARK)
c.alignment = center()
for col in "BCDEFGH":
    dash[f"{col}2"].fill = hfill(C_DARK)

# ── KPI cards row ──────────────────────────────────────────────
latest  = master.iloc[-1]
prev    = master.iloc[-2]

kpis = [
    ("WEEKLY REVENUE",   f"₹{latest['revenue']/1e6:.1f}M",
     f"{'▲' if latest['revenue']>prev['revenue'] else '▼'} {abs((latest['revenue']-prev['revenue'])/prev['revenue']*100):.1f}% WoW",
     C_ORANGE),
    ("GROSS MARGIN",     f"{latest['gpm_pct']:.1f}%",
     f"Target: 38–42%",
     C_GREEN if 38 <= latest['gpm_pct'] <= 42 else C_ORANGE),
    ("EBITDA MARGIN",    f"{latest['ebitda_pct']:.1f}%",
     f"{'▲' if latest['ebitda_pct']>prev['ebitda_pct'] else '▼'} {abs(latest['ebitda_pct']-prev['ebitda_pct']):.1f}pp WoW",
     C_BLUE),
    ("VS BUDGET",        f"{latest['budget_var_pct']:+.1f}%",
     "Above plan ✓" if latest['budget_var_pct'] >= 0 else "Below plan ✗",
     C_GREEN if latest['budget_var_pct'] >= 0 else C_ORANGE),
    ("COST / ORDER",     f"₹{latest['cpo']:.1f}",
     "✓ Within limit" if latest['cpo'] <= 52 else "⚠ Above ₹52 limit",
     C_GREEN if latest['cpo'] <= 52 else C_ORANGE),
    ("TOTAL ORDERS",     f"{int(latest['total_orders']):,}",
     f"{'▲' if latest['total_orders']>prev['total_orders'] else '▼'} WoW",
     C_BLUE),
]

kpi_cols = ["B","C","D","E","F","G"]
for i, (label, value, sub, color) in enumerate(kpis):
    col = kpi_cols[i]
    row_lbl, row_val, row_sub = 4, 5, 6
    dash.row_dimensions[4].height = 16
    dash.row_dimensions[5].height = 28
    dash.row_dimensions[6].height = 16
    dash.row_dimensions[7].height = 8

    for r in [4,5,6]:
        dash[f"{col}{r}"].fill = hfill(C_LIGHT_BG)
        dash[f"{col}{r}"].border = Border(
            left  =side(color) if i==0 else side(),
            right =side(color) if i==5 else side(),
            top   =side(color) if r==4 else side(),
            bottom=side(color) if r==6 else side(),
        )

    cl = dash[f"{col}4"]
    cl.value = label
    cl.font  = Font(name="Arial", size=8, bold=True, color="7A7268")
    cl.alignment = center()

    cv = dash[f"{col}5"]
    cv.value = value
    cv.font  = Font(name="Arial", size=17, bold=True, color=color)
    cv.alignment = center()

    cs = dash[f"{col}6"]
    cs.value = sub
    cs.font  = Font(name="Arial", size=8, italic=True,
                    color=C_GREEN if "✓" in sub or "▲" in sub else
                          C_ORANGE if "✗" in sub or "⚠" in sub or "▼" in sub else "7A7268")
    cs.alignment = center()

# ── Section header helper ──────────────────────────────────────
def section_header(sheet, row, text):
    sheet.merge_cells(f"B{row}:H{row}")
    c = sheet[f"B{row}"]
    c.value = text
    c.font  = Font(name="Arial", size=10, bold=True, color=C_WHITE)
    c.fill  = hfill(C_DARK)
    c.alignment = left()
    c.border = border()
    sheet.row_dimensions[row].height = 22

def col_header(sheet, row, cols_labels, bg=C_BLUE):
    for col, label in cols_labels.items():
        c = sheet[f"{col}{row}"]
        c.value = label
        c.font  = Font(name="Arial", size=9, bold=True, color=C_WHITE)
        c.fill  = hfill(bg)
        c.alignment = center()
        c.border = border()

# ── Weekly KPI table ──────────────────────────────────────────
START_ROW = 9
section_header(dash, START_ROW, "  WEEKLY KPI SUMMARY  (All 26 Weeks)")
dash.row_dimensions[START_ROW].height = 22

hdrs = {"B":"Week","C":"Revenue (₹M)","D":"vs Budget %","E":"GPM %",
        "F":"EBITDA %","G":"Cost/Order ₹","H":"Orders"}
col_header(dash, START_ROW+1, hdrs)

for i, (_, row) in enumerate(master.iterrows()):
    r   = START_ROW + 2 + i
    bg  = C_WHITE if i % 2 == 0 else "F7F4EF"
    dash.row_dimensions[r].height = 16
    vals = {
        "B": f"W{int(row['week_num'])}  {row['week_start']}",
        "C": f"₹{row['revenue']/1e6:.2f}M",
        "D": f"{row['budget_var_pct']:+.2f}%",
        "E": f"{row['gpm_pct']:.1f}%",
        "F": f"{row['ebitda_pct']:.1f}%",
        "G": f"₹{row['cpo']:.1f}",
        "H": f"{int(row['total_orders']):,}",
    }
    for col, val in vals.items():
        c = dash[f"{col}{r}"]
        c.value = val
        c.font  = Font(name="Arial", size=9, color=C_DARK)
        c.fill  = hfill(bg)
        c.alignment = center()
        c.border = border()
        # Color-code budget variance
        if col == "D":
            var = row['budget_var_pct']
            c.fill = hfill(C_GRN_LIGHT if var >= 0 else C_RED_LIGHT)
            c.font = Font(name="Arial", size=9, bold=True,
                          color=C_GREEN if var >= 0 else C_ORANGE)
        if col == "G":
            cpo = row['cpo']
            c.fill = hfill(C_RED_LIGHT if cpo > 52 else C_GRN_LIGHT)

# ════════════════════════════════════════════════════════════════
# SHEET 2 — CHARTS DATA + CHARTS
# ════════════════════════════════════════════════════════════════
ch_sheet = wb.create_sheet("📈 Charts")
ch_sheet.sheet_view.showGridLines = False
for r in range(1,100): ch_sheet.row_dimensions[r].height = 15
for c,w in [("A",3),("B",14),("C",14),("D",14),("E",14),("F",14),("G",14),("H",14),("I",3)]:
    ch_sheet.column_dimensions[c].width = w

# Write chart data
section_header(ch_sheet, 1, "  CHART DATA — Revenue vs Budget (26 Weeks)")
col_header(ch_sheet, 2, {"B":"Week","C":"Actual Rev ₹M","D":"Budget ₹M","E":"WoW Growth %"})

for i, (_, row) in enumerate(master.iterrows()):
    r = 3 + i
    wow = "" if i == 0 else f"{((master.iloc[i]['revenue']-master.iloc[i-1]['revenue'])/master.iloc[i-1]['revenue']*100):.2f}"
    ch_sheet[f"B{r}"] = f"W{int(row['week_num'])}"
    ch_sheet[f"C{r}"] = round(row["revenue"]/1e6, 2)
    ch_sheet[f"D{r}"] = round(row["budget"]/1e6, 2)
    ch_sheet[f"E{r}"] = float(wow) if wow else None
    for col in "BCDE":
        ch_sheet[f"{col}{r}"].font   = Font(name="Arial", size=9)
        ch_sheet[f"{col}{r}"].border = border()
        ch_sheet[f"{col}{r}"].alignment = center()
        bg = C_WHITE if i%2==0 else "F7F4EF"
        ch_sheet[f"{col}{r}"].fill = hfill(bg)

# Revenue vs Budget bar chart
bar = BarChart()
bar.type    = "col"
bar.grouping = "clustered"
bar.title   = "Revenue vs Budget — 26-Week Trend"
bar.y_axis.title = "Revenue (₹M)"
bar.x_axis.title = "Week"
bar.style   = 10
bar.width   = 22
bar.height  = 12

actual_data = Reference(ch_sheet, min_col=3, max_col=3, min_row=2, max_row=2+len(master))
budget_data = Reference(ch_sheet, min_col=4, max_col=4, min_row=2, max_row=2+len(master))
weeks_ref   = Reference(ch_sheet, min_col=2, min_row=3, max_row=2+len(master))
bar.add_data(actual_data, titles_from_data=True)
bar.add_data(budget_data, titles_from_data=True)
bar.set_categories(weeks_ref)
bar.series[0].graphicalProperties.solidFill = C_ORANGE
bar.series[1].graphicalProperties.solidFill = C_BLUE
ch_sheet.add_chart(bar, "B31")

# GPM trend line chart
gpm_row_start = 31
section_header(ch_sheet, gpm_row_start, "  GPM % Trend Data")
col_header(ch_sheet, gpm_row_start+1, {"B":"Week","C":"GPM %","D":"EBITDA %","E":"OPEX %"})
for i, (_, row) in enumerate(master.iterrows()):
    r = gpm_row_start + 2 + i
    ch_sheet[f"B{r}"] = f"W{int(row['week_num'])}"
    ch_sheet[f"C{r}"] = row["gpm_pct"]
    ch_sheet[f"D{r}"] = row["ebitda_pct"]
    ch_sheet[f"E{r}"] = row["opex_pct"]
    for col in "BCDE":
        ch_sheet[f"{col}{r}"].font  = Font(name="Arial", size=9)
        ch_sheet[f"{col}{r}"].border = border()
        ch_sheet[f"{col}{r}"].alignment = center()

line = LineChart()
line.title  = "Margin Trends — GPM / EBITDA / OPEX %"
line.y_axis.title = "Margin %"
line.x_axis.title = "Week"
line.style  = 10
line.width  = 22
line.height = 12
for col_i, color, label in [(3,C_GREEN,"GPM %"),(4,C_BLUE,"EBITDA %"),(5,C_ORANGE,"OPEX %")]:
    d = Reference(ch_sheet, min_col=col_i, min_row=gpm_row_start+1, max_row=gpm_row_start+1+len(master))
    line.add_data(d, titles_from_data=True)
line.set_categories(Reference(ch_sheet, min_col=2, min_row=gpm_row_start+2, max_row=gpm_row_start+1+len(master)))
line.series[0].graphicalProperties.line.solidFill = C_GREEN
line.series[1].graphicalProperties.line.solidFill = C_BLUE
line.series[2].graphicalProperties.line.solidFill = C_ORANGE
ch_sheet.add_chart(line, "K31")

# ════════════════════════════════════════════════════════════════
# SHEET 3 — REGION SCORECARD
# ════════════════════════════════════════════════════════════════
reg = wb.create_sheet("🗺️ Region Scorecard")
reg.sheet_view.showGridLines = False
for c,w in [("A",3),("B",16),("C",16),("D",14),("E",14),("F",16),("G",14),("H",14),("I",3)]:
    reg.column_dimensions[c].width = w

section_header(reg, 1, "  REGION PERFORMANCE SCORECARD  (All-Period Average)")
col_header(reg, 2, {
    "B":"Region","C":"Total Revenue","D":"GPM %","E":"EBITDA %",
    "F":"Total Orders","G":"Cost/Order ₹","H":"vs Budget %"
})

for i, (_, row) in enumerate(region_df.iterrows()):
    r  = 3 + i
    bg = C_WHITE if i%2==0 else "F7F4EF"
    reg.row_dimensions[r].height = 20
    vals = {
        "B": row["region"],
        "C": f"₹{row['revenue']/1e6:.1f}M",
        "D": f"{row['gpm_pct']:.1f}%",
        "E": f"{row['ebitda_pct']:.1f}%",
        "F": f"{int(row['orders']):,}",
        "G": f"₹{row['cost_per_order']:.1f}",
        "H": f"{row['budget_var']:+.1f}%",
    }
    for col, val in vals.items():
        c = reg[f"{col}{r}"]
        c.value = val
        c.font  = Font(name="Arial", size=10, color=C_DARK,
                       bold=(col=="B"))
        c.fill  = hfill(bg)
        c.alignment = center()
        c.border = border()
        if col == "H":
            v = row["budget_var"]
            c.fill = hfill(C_GRN_LIGHT if v >= 0 else C_RED_LIGHT)
            c.font = Font(name="Arial", size=10, bold=True,
                          color=C_GREEN if v >= 0 else C_ORANGE)
        if col == "G":
            c.fill = hfill(C_RED_LIGHT if row["cost_per_order"] > 52 else C_GRN_LIGHT)

# ════════════════════════════════════════════════════════════════
# SHEET 4 — MANAGEMENT COMMENTARY
# ════════════════════════════════════════════════════════════════
comm = wb.create_sheet("📋 Commentary")
comm.sheet_view.showGridLines = False
comm.column_dimensions["A"].width = 3
comm.column_dimensions["B"].width = 22
comm.column_dimensions["C"].width = 60
comm.column_dimensions["D"].width = 45

section_header(comm, 1, f"  MANAGEMENT COMMENTARY  ·  Latest Week: W{int(latest['week_num'])}  ({latest['week_start']})")
col_header(comm, 2, {"B":"KPI","C":"Observation","D":"Recommended Action"}, bg=C_BLUE)

# Auto-generate commentary from actual data
commentaries = []
bv = latest['budget_var_pct']
commentaries.append((
    "Revenue vs Budget",
    f"Revenue {'beat' if bv>=0 else 'missed'} budget by {abs(bv):.1f}%. "
    f"Actual: ₹{latest['revenue']/1e6:.2f}M vs Plan: ₹{latest['budget']/1e6:.2f}M.",
    "Continue current momentum." if bv >= 0
    else "Investigate demand shortfall. Review promotional activity in underperforming regions.",
    bv >= 0
))
gpm = latest['gpm_pct']
commentaries.append((
    "Gross Profit Margin",
    f"GPM at {gpm:.1f}% — {'within' if 38<=gpm<=42 else 'outside'} target band of 38–42%. "
    f"Gross profit: ₹{latest['gross_profit']/1e6:.2f}M.",
    "Maintain pricing discipline." if 38<=gpm<=42
    else ("Review COGS drivers. Engage Procurement on input cost reduction." if gpm<38
          else "Validate pricing — margin above band may indicate underinvestment in growth."),
    38<=gpm<=42
))
cpo = latest['cpo']
commentaries.append((
    "Cost per Order",
    f"Fulfillment cost at ₹{cpo:.1f} per order. "
    f"Threshold: ₹52. Total fulfillment spend: ₹{latest['fulfillment_cost']/1e6:.2f}M.",
    "Monitor carrier mix for efficiency gains." if cpo <= 52
    else "Escalate to Logistics team. Review last-mile carrier allocation by region.",
    cpo <= 52
))
ebitda = latest['ebitda_pct']
commentaries.append((
    "EBITDA Margin",
    f"EBITDA margin at {ebitda:.1f}%. EBITDA: ₹{latest['ebitda']/1e6:.2f}M. "
    f"OPEX running at {latest['opex_pct']:.1f}% of revenue.",
    "Strong profitability. Evaluate reinvestment opportunities." if ebitda >= 15
    else "Review OPEX line items. Identify non-essential spend to reduce below 20%.",
    ebitda >= 15
))
top_ret = cat_df.iloc[0]
commentaries.append((
    "Category Returns",
    f"{top_ret['category']} has highest return rate at {top_ret['return_rate_pct']:.1f}%. "
    f"Returns erode net revenue and inflate fulfillment cost.",
    f"Engage {top_ret['category']} category team to review product quality and size/fit guidance.",
    top_ret['return_rate_pct'] <= 10
))

for i, (kpi, obs, action, good) in enumerate(commentaries):
    r  = 3 + i
    bg = C_WHITE if i%2==0 else "F7F4EF"
    comm.row_dimensions[r].height = 36

    c1 = comm[f"B{r}"]
    c1.value = kpi
    c1.font  = Font(name="Arial", size=9, bold=True, color=C_DARK)
    c1.fill  = hfill(bg)
    c1.alignment = left()
    c1.border = border()

    c2 = comm[f"C{r}"]
    c2.value = obs
    c2.font  = Font(name="Arial", size=9, color=C_DARK)
    c2.fill  = hfill(bg)
    c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c2.border = border()

    c3 = comm[f"D{r}"]
    c3.value = f"→  {action}"
    c3.font  = Font(name="Arial", size=9, bold=True,
                    color=C_GREEN if good else C_ORANGE)
    c3.fill  = hfill(C_GRN_LIGHT if good else C_YLW_LIGHT)
    c3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    c3.border = border()

# ════════════════════════════════════════════════════════════════
# SHEET 5 — CATEGORY RETURNS
# ════════════════════════════════════════════════════════════════
cat_sh = wb.create_sheet("📦 Category Analysis")
cat_sh.sheet_view.showGridLines = False
for c,w in [("A",3),("B",20),("C",16),("D",16),("E",18),("F",18),("G",3)]:
    cat_sh.column_dimensions[c].width = w

section_header(cat_sh, 1, "  PRODUCT CATEGORY — Returns & Revenue Analysis")
col_header(cat_sh, 2, {
    "B":"Category","C":"Total Orders","D":"Returns","E":"Return Rate %","F":"Total Revenue"
})

for i, (_, row) in enumerate(cat_df.iterrows()):
    r  = 3 + i
    bg = C_WHITE if i%2==0 else "F7F4EF"
    cat_sh.row_dimensions[r].height = 20
    vals = {
        "B": row["category"],
        "C": f"{int(row['orders']):,}",
        "D": f"{int(row['returns']):,}",
        "E": f"{row['return_rate_pct']:.2f}%",
        "F": f"₹{row['revenue']/1e6:.1f}M",
    }
    for col, val in vals.items():
        c = cat_sh[f"{col}{r}"]
        c.value = val
        c.font  = Font(name="Arial", size=10, color=C_DARK, bold=(col=="B"))
        c.fill  = hfill(bg)
        c.alignment = center()
        c.border = border()
        if col == "E":
            rate = row["return_rate_pct"]
            c.fill = hfill(C_RED_LIGHT if rate > 10 else
                           C_YLW_LIGHT if rate > 5  else C_GRN_LIGHT)
            c.font = Font(name="Arial", size=10, bold=True,
                          color=C_ORANGE if rate > 10 else
                                C_YELLOW  if rate > 5  else C_GREEN)

# ── Save ──────────────────────────────────────────────────────
out = "weekly_kpi_report.xlsx"
wb.save(out)
print(f"\n✅  Saved: {out}")
print("   Sheets:")
for s in wb.sheetnames:
    print(f"     {s}")
print("\n🚀  Open weekly_kpi_report.xlsx — your dashboard is ready!")
print("   Best starting sheet: '📊 Dashboard'")